"""
tools/export_mark_scheme_review.py
====================================
PHASE: build

Converts the {subject}_mark_scheme_review.csv produced by extract_mark_scheme.py
into a formatted Excel workbook for manual verification.

Usage:
    python tools/export_mark_scheme_review.py --subject Economics
"""

import argparse
import csv
import os
import sys
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).resolve().parents[1] / ".env")

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


# Conditional-format fill colours — priority order documented alongside each:
FILL_RED    = PatternFill("solid", fgColor="FF9999")   # parser_artifact=1  (excluded by design, not a gap)
FILL_ORANGE = PatternFill("solid", fgColor="FFB366")   # needs_manual_entry=1 or [REVIEW NEEDED:]
FILL_MAROON = PatternFill("solid", fgColor="C00000")   # genuine unmapped gap (empty mapped_objective_id, not artifact, not manual)
FILL_YELLOW = PatternFill("solid", fgColor="FFFF99")   # verified=0 (routine unverified row)
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")   # header row

HEADER_FONT  = Font(bold=True, color="FFFFFF")
MAROON_FONT  = Font(color="FFFFFF")                    # white text on dark-red background
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")


def export(subject: str, reports_root: str) -> Path:
    csv_path  = Path(reports_root) / f"{subject}_mark_scheme_review.csv"
    xlsx_path = Path(reports_root) / f"{subject}_mark_scheme_review.xlsx"

    if not csv_path.exists():
        sys.exit(f"ERROR: CSV not found: {csv_path}\nRun extract_mark_scheme.py first.")

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows   = list(reader)
        cols   = reader.fieldnames or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = subject[:31]  # Excel sheet name limit

    # ── Header row ───────────────────────────────────────────────────────────
    ws.append(cols)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = FILL_HEADER
        cell.alignment = WRAP_ALIGN

    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_dict in rows:
        row_values = [row_dict.get(c, "") for c in cols]
        ws.append(row_values)

        excel_row = ws.max_row
        point_text      = row_dict.get("point_text", "")
        verified        = str(row_dict.get("verified", "0")).strip()
        mapped_obj      = str(row_dict.get("mapped_objective_id", "")).strip()
        parser_artifact = str(row_dict.get("parser_artifact", "0")).strip()
        needs_manual    = str(row_dict.get("needs_manual_entry", "0")).strip()

        # Priority: artifact > manual/incomplete > genuine gap > unverified
        if parser_artifact == "1":
            fill     = FILL_RED     # intentional exclusion, not a content gap
            row_font = None
        elif needs_manual == "1" or "[REVIEW NEEDED:" in point_text:
            fill     = FILL_ORANGE  # requires manual entry against source PDF
            row_font = None
        elif not mapped_obj:
            fill     = FILL_MAROON  # genuine unmapped gap — needs investigation
            row_font = MAROON_FONT
        elif verified == "0":
            fill     = FILL_YELLOW  # unverified but mapped
            row_font = None
        else:
            fill     = None
            row_font = None

        if fill:
            for cell in ws[excel_row]:
                cell.fill = fill
                if row_font:
                    cell.font = row_font

        # Wrap text on all cells
        for cell in ws[excel_row]:
            cell.alignment = WRAP_ALIGN

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = {
        "question_num":       10,
        "question_part":      14,
        "so_codes":           14,
        "point_text":         60,
        "marks_value":        12,
        "point_order":        12,
        "profile":            10,
        "source_page":        12,
        "raw_excerpt":        50,
        "mapped_objective_id": 24,
        "verified":           10,
    }
    for col_idx, col_name in enumerate(cols, 1):
        width = col_widths.get(col_name, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(xlsx_path)
    return xlsx_path


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Export mark-scheme review CSV to formatted Excel."
    )
    ap.add_argument("--subject", required=True, help="Subject ID (e.g. Economics)")
    args = ap.parse_args()

    reports_root = os.getenv("REPORTS_ROOT")
    if not reports_root:
        sys.exit("ERROR: REPORTS_ROOT not set in .env")

    out = export(args.subject, reports_root)
    print(f"Excel written: {out}")


if __name__ == "__main__":
    main()
