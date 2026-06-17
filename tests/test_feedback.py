"""
tests/test_feedback.py
======================
Stage 12 (Build Playbook v3.1) tests for the feedback loop.

Real in-memory SQLite (schema.sql + apply_runtime_migrations + sqlite-vec, FKs
ON) so the user_feedback CHECK constraints and FKs are genuinely exercised. The
endpoint is driven through a TestClient; the report is called directly with a
tempdir as REPORTS_ROOT. No Ollama, no SSD.

  1. POST valid feedback -> 200, ok=True, one row written.
  2. POST sentiment='angry'          -> 422 (Pydantic enum).
  3. POST feedback_type='something'  -> 422 (Pydantic enum).
  4. POST unknown objective_id       -> 400 (FK), ok=False.
  5. Report groups + orders correctly and writes a formatted Excel file.
  6. Empty feedback writes a header-only file.

Run: pytest tests/test_feedback.py -v
"""

import sqlite3
import sys
from pathlib import Path

import pytest
from starlette.testclient import TestClient

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "backend"))

SCHEMA_PATH = ROOT / "backend" / "db" / "schema.sql"

import app as app_module          # noqa: E402  (apply_runtime_migrations + the app)
import feedback_report            # noqa: E402

SUBJECT = "Principles_of_Business"


# --- in-memory DB ----------------------------------------------------------
def open_test_db() -> sqlite3.Connection:
    try:
        import sqlite_vec
    except ImportError:
        pytest.skip("sqlite-vec not installed -- skipping feedback tests")
    # check_same_thread=False: the FastAPI TestClient runs sync endpoints in a
    # worker thread, so the connection must be usable across threads.
    db = sqlite3.connect(":memory:", check_same_thread=False)
    db.enable_load_extension(True)
    sqlite_vec.load(db)
    db.enable_load_extension(False)
    db.execute("PRAGMA foreign_keys = ON")
    db.row_factory = sqlite3.Row
    for stmt in SCHEMA_PATH.read_text(encoding="utf-8").split(";"):
        if stmt.strip():
            db.execute(stmt)
    db.commit()
    app_module.apply_runtime_migrations(db)  # creates user_feedback + indexes
    return db


def seed(db: sqlite3.Connection, objective_ids=("POB-1.1",)) -> None:
    """Locked subject + one section + the named objectives (FK targets)."""
    db.execute(
        "INSERT INTO subjects (subject_id, display_name, syllabus_locked) VALUES (?, ?, 1)",
        (SUBJECT, "Principles of Business"),
    )
    db.execute(
        "INSERT INTO syllabus_sections (section_id, subject_id, title, section_num) "
        "VALUES (?, ?, ?, ?)",
        ("POB-SEC-1", SUBJECT, "Nature of Business", "1"),
    )
    for i, oid in enumerate(objective_ids, 1):
        db.execute(
            "INSERT INTO objectives (objective_id, section_id, subject_id, objective_num, "
            "content_stmt) VALUES (?, ?, ?, ?, ?)",
            (oid, "POB-SEC-1", SUBJECT, f"1.{i}",
             f"Explain concept number {i} of business operations and management"),
        )
    db.commit()


@pytest.fixture
def db():
    conn = open_test_db()
    seed(conn)
    yield conn
    conn.close()


@pytest.fixture
def client(db):
    """TestClient with app.state.db pointed at the seeded in-memory DB.

    No `with` block -> lifespan (SSD check + real DB open) does NOT run; we own
    app.state.db here, exactly like tests/test_api.py.
    """
    app_module.app.state.db = db
    return TestClient(app_module.app)


def feedback_count(db, objective_id="POB-1.1") -> int:
    return db.execute(
        "SELECT COUNT(*) FROM user_feedback WHERE objective_id = ?", (objective_id,)
    ).fetchone()[0]


# --- endpoint tests --------------------------------------------------------
def test_post_valid_feedback_writes_a_row(client, db):
    res = client.post("/api/feedback", json={
        "objective_id": "POB-1.1",
        "subject_id": SUBJECT,
        "feedback_type": "lesson",
        "sentiment": "positive",
    })
    assert res.status_code == 200
    body = res.json()
    assert body["ok"] is True
    assert isinstance(body["feedback_id"], int)

    assert feedback_count(db) == 1
    row = db.execute(
        "SELECT feedback_type, sentiment FROM user_feedback WHERE objective_id = ?",
        ("POB-1.1",),
    ).fetchone()
    assert row["feedback_type"] == "lesson"
    assert row["sentiment"] == "positive"


def test_post_invalid_sentiment_returns_422(client):
    res = client.post("/api/feedback", json={
        "objective_id": "POB-1.1",
        "subject_id": SUBJECT,
        "feedback_type": "lesson",
        "sentiment": "angry",  # not in the Literal enum
    })
    assert res.status_code == 422


def test_post_invalid_feedback_type_returns_422(client):
    res = client.post("/api/feedback", json={
        "objective_id": "POB-1.1",
        "subject_id": SUBJECT,
        "feedback_type": "something_else",  # not in the Literal enum
        "sentiment": "positive",
    })
    assert res.status_code == 422


def test_post_unknown_objective_returns_400(client, db):
    res = client.post("/api/feedback", json={
        "objective_id": "POB-999.99",  # no such objective -> FK violation
        "subject_id": SUBJECT,
        "feedback_type": "lesson",
        "sentiment": "negative",
    })
    assert res.status_code == 400
    body = res.json()
    assert body["ok"] is False
    assert "objective" in body["error"].lower() or "subject" in body["error"].lower()
    assert feedback_count(db, "POB-999.99") == 0  # nothing written


# --- report tests ----------------------------------------------------------
def _add_feedback(db, objective_id, sentiment, n):
    for _ in range(n):
        db.execute(
            "INSERT INTO user_feedback (objective_id, subject_id, feedback_type, sentiment) "
            "VALUES (?, ?, 'lesson', ?)",
            (objective_id, SUBJECT, sentiment),
        )
    db.commit()


def test_report_groups_orders_and_writes_excel(tmp_path):
    from openpyxl import load_workbook

    conn = open_test_db()
    seed(conn, objective_ids=("POB-1.1", "POB-1.2", "POB-1.3"))
    try:
        # obj A (POB-1.1): 2 negative + 1 positive -> neg+confused=2, total=3
        _add_feedback(conn, "POB-1.1", "negative", 2)
        _add_feedback(conn, "POB-1.1", "positive", 1)
        # obj B (POB-1.2): 1 confused + 1 negative -> neg+confused=2, total=2
        _add_feedback(conn, "POB-1.2", "confused", 1)
        _add_feedback(conn, "POB-1.2", "negative", 1)
        # obj C (POB-1.3): 2 positive -> neg+confused=0 -> excluded by HAVING
        _add_feedback(conn, "POB-1.3", "positive", 2)

        out_path, count = feedback_report.generate_report(
            conn, SUBJECT, str(tmp_path), today="2026-06-17",
        )
    finally:
        conn.close()

    assert out_path.exists()
    assert count == 2, "only the two flagged objectives are reported"

    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.title == "Top objectives for review"
    # A and B tie on neg+confused=2; total_feedback DESC breaks it -> A (3) before B (2).
    assert ws.cell(row=2, column=1).value == "POB-1.1"
    assert ws.cell(row=3, column=1).value == "POB-1.2"
    assert ws.max_row == 3, "obj C (no negative/confused) is NOT in the report"
    ids_in_sheet = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert "POB-1.3" not in ids_in_sheet
    # header bold + frozen
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.freeze_panes == "A2"


def test_empty_feedback_writes_header_only_file(tmp_path):
    from openpyxl import load_workbook

    conn = open_test_db()
    seed(conn)  # subject + objective, but NO user_feedback rows
    try:
        out_path, count = feedback_report.generate_report(
            conn, SUBJECT, str(tmp_path), today="2026-06-17",
        )
    finally:
        conn.close()

    assert out_path.exists()
    assert count == 0

    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Objective ID"  # header present
    # No data rows: max_row is 1 (the header). Check this BEFORE touching row 2 --
    # ws.cell(2, 1) would materialise a phantom cell and bump max_row to 2.
    assert ws.max_row == 1
