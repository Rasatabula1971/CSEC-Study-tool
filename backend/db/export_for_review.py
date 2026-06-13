"""
backend/db/export_for_review.py
===============================
Exports a subject's objectives to an Excel workbook for manual syllabus sign-off:
    {REPORTS_ROOT}\\{subject}_syllabus_review.xlsx

The reviewer reads each objective against the real CXC PDF, ticks the Approved
column, and only then is lock_subject.py run. Header row is bold and frozen.

Usage:
    python backend/db/export_for_review.py --subject Principles_of_Business
"""

import argparse
import os
import sqlite3
import sys
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).resolve().parents[2] / ".env")

# (DB column / expression, spreadsheet header, column width)
COLUMNS = [
    ("o.objective_id",   "objective_id",   16),
    ("s.section_num",    "section_num",    12),
    ("s.title",          "section_title",  34),
    ("o.objective_num",  "objective_num",  14),
    ("o.content_stmt",   "content_stmt",   70),
    ("o.skill_type",     "skill_type",     16),
    ("o.command_words",  "command_words",  20),
    ("o.exam_weight",    "exam_weight",    12),
    ("o.verified",       "verified",       10),
]


def open_db(db_path: str) -> sqlite3.Connection:
    try:
        import sqlite_vec
    except ImportError:
        sys.exit("ERROR: sqlite-vec is not installed. Run: pip install sqlite-vec")
    db = sqlite3.connect(db_path)
    db.enable_load_extension(True)
    sqlite_vec.load(db)
    db.enable_load_extension(False)
    db.execute("PRAGMA foreign_keys = ON")
    db.row_factory = sqlite3.Row
    return db


def fetch_objectives(db: sqlite3.Connection, subject_id: str) -> list[sqlite3.Row]:
    select = ", ".join(col for col, _, _ in COLUMNS)
    return db.execute(
        f"""
        SELECT {select}
        FROM   objectives o
        JOIN   syllabus_sections s ON s.section_id = o.section_id
        WHERE  o.subject_id = ?
        ORDER  BY CAST(s.section_num AS INTEGER), o.objective_num
        """,
        (subject_id,),
    ).fetchall()


def build_workbook(rows: list[sqlite3.Row], subject_id: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Objectives"

    headers = [hdr for _, hdr, _ in COLUMNS] + ["Approved (Y/N)", "Reviewer notes"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    wrap = Alignment(vertical="top", wrap_text=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append([row[hdr] for _, hdr, _ in COLUMNS] + ["", ""])

    # Column widths + wrap on the body.
    widths = [w for _, _, w in COLUMNS] + [16, 40]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"  # freeze the header row
    return wb


def main() -> None:
    ap = argparse.ArgumentParser(description="Export objectives to an Excel review file.")
    ap.add_argument("--subject", required=True, help="e.g. Principles_of_Business")
    args = ap.parse_args()

    db_path = os.getenv("DB_PATH")
    reports_root = os.getenv("REPORTS_ROOT")
    if not db_path:
        sys.exit("ERROR: DB_PATH not set in .env")
    if not reports_root:
        sys.exit("ERROR: REPORTS_ROOT not set in .env")
    if not Path(db_path).exists():
        sys.exit(f"ERROR: database not found at {db_path}. Run init_db.py first.")

    db = open_db(db_path)
    try:
        rows = fetch_objectives(db, args.subject)
    finally:
        db.close()

    if not rows:
        sys.exit(
            f"ERROR: no objectives found for subject '{args.subject}'.\n"
            "Run syllabus_parser.py first (and check the --subject name)."
        )

    out_dir = Path(reports_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{args.subject}_syllabus_review.xlsx"

    wb = build_workbook(rows, args.subject)
    wb.save(out_path)

    print(f"Objectives exported: {len(rows)}")
    print(f"Review file written: {out_path}")
    print("Open it, verify every row against the CXC PDF, then run lock_subject.py.")


if __name__ == "__main__":
    main()
