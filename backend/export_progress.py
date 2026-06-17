# PHASE: build
"""
backend/export_progress.py
==========================
Parent-facing study-progress export. Produces a colour-coded Excel workbook so a
parent can see, at a glance, which objectives are mastered, which need work, and
which the student hasn't started -- without ever touching the database.

Everything here is deterministic SQLite + openpyxl (CLAUDE.md "Rule 2"): the
status of each objective comes straight from study_plan / weakness_log, never the
LLM. Objectives are LEFT-joined to both tables, so an objective with no study_plan
and no weakness_log row shows as "Not started".

Columns:
    Section | Objective | Status | Last Score | Leitner Box | Next Review | Times Passed

Status colours:
    Mastered    -> green        (study_plan.status = 'mastered')
    Met once    -> light green  (study_plan.status = 'met_once')
    Needs work  -> orange       (unmet/in_progress but has a weakness_log entry)
    Not started -> no fill      (no study_plan progress and no weakness_log entry)

Usage:
    python backend/export_progress.py --subject Principles_of_Business
"""

import argparse
import os
import sqlite3
import sys
from datetime import date
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).resolve().parents[1] / ".env")

HEADERS = [
    "Section", "Objective", "Status", "Last Score",
    "Leitner Box", "Next Review", "Times Passed",
]
STATUS_COL = 3  # 1-based column index of the Status column (for fill + width).

# Fills, deliberately distinct so a parent can tell mastered from met-once at a glance.
GREEN = "63BE7B"        # mastered
LIGHT_GREEN = "C6EFCE"  # met once
ORANGE = "FFC000"       # unmet but attempted (has a weakness record)
HEADER_FILL = "305496"


def open_db(db_path: str) -> sqlite3.Connection:
    try:
        import sqlite_vec
    except ImportError:
        sys.exit("ERROR: sqlite-vec is not installed. Run: pip install sqlite-vec")
    db = sqlite3.connect(db_path)
    db.enable_load_extension(True)
    sqlite_vec.load(db)
    db.enable_load_extension(False)
    db.execute("PRAGMA foreign_keys = ON")
    db.row_factory = sqlite3.Row
    return db


def fetch_progress(db: sqlite3.Connection, subject_id: str) -> list[sqlite3.Row]:
    """Every objective for the subject, LEFT-joined to study_plan + weakness_log.

    LEFT joins keep objectives that have no progress yet (status / score_pct etc.
    come back NULL -> rendered as "Not started"). Ordered in syllabus order so the
    sheet reads top-to-bottom like the syllabus.
    """
    return db.execute(
        """
        SELECT o.objective_id,
               o.objective_num,
               o.content_stmt,
               s.title       AS section_title,
               s.section_num AS section_num,
               p.status      AS status,
               p.met_count   AS met_count,
               w.score_pct   AS score_pct,
               w.leitner_box AS leitner_box,
               w.next_review AS next_review
        FROM   objectives o
        JOIN   syllabus_sections s ON s.section_id = o.section_id
        LEFT   JOIN study_plan   p ON p.objective_id = o.objective_id
                                  AND p.subject_id   = o.subject_id
        LEFT   JOIN weakness_log w ON w.objective_id = o.objective_id
                                  AND w.subject_id   = o.subject_id
        WHERE  o.subject_id = ?
        ORDER  BY CAST(s.section_num AS INTEGER), o.objective_num
        """,
        (subject_id,),
    ).fetchall()


def classify_status(status: str | None, has_weakness: bool) -> tuple[str, str | None]:
    """Map (study_plan.status, weakness-present?) to a label and a fill colour.

    Returns (label, fgColor) where fgColor is None for "Not started" (no fill).
    """
    if status == "mastered":
        return "Mastered", GREEN
    if status == "met_once":
        return "Met once", LIGHT_GREEN
    if has_weakness:  # unmet / in_progress / no plan row, but attempted at least once
        return "Needs work", ORANGE
    return "Not started", None


def _summary_text(rows: list[sqlite3.Row]) -> tuple[str, int, int, int]:
    """Build the top summary line. Returns (text, mastered, total, percent)."""
    total = len(rows)
    mastered = sum(1 for r in rows if r["status"] == "mastered")
    percent = round(100 * mastered / total) if total else 0
    return f"Mastered: {mastered}/{total} ({percent}%)", mastered, total, percent


def build_workbook(rows: list[sqlite3.Row], subject_id: str):
    """Assemble the colour-coded workbook. Pure openpyxl -- no DB, no file I/O.

    Layout: row 1 = bold summary, row 2 = header, rows 3+ = one per objective.
    Rows 1-2 are frozen so the summary and header stay visible while scrolling.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Progress"

    # Row 1: summary, bold.
    summary, _, _, _ = _summary_text(rows)
    summary_cell = ws.cell(row=1, column=1, value=summary)
    summary_cell.font = Font(bold=True, size=12)

    # Row 2: header.
    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # Rows 3+: one per objective.
    for r in rows:
        has_weakness = r["score_pct"] is not None
        label, fg = classify_status(r["status"], has_weakness)
        ws.append([
            r["section_title"],
            f'{r["objective_num"]} {r["content_stmt"]}',
            label,
            f'{r["score_pct"]}%' if r["score_pct"] is not None else "",
            r["leitner_box"] if r["leitner_box"] is not None else "",
            r["next_review"] or "",
            r["met_count"] if r["met_count"] is not None else 0,
        ])
        if fg is not None:
            ws.cell(row=ws.max_row, column=STATUS_COL).fill = PatternFill("solid", fgColor=fg)

    # Wrap the long Objective column; keep the rest top-aligned.
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 2))

    _auto_width(ws)
    ws.freeze_panes = "A3"  # freeze the summary + header rows
    return wb


def _auto_width(ws) -> None:
    """Size each column to its widest cell (capped so content_stmt stays readable)."""
    caps = {2: 70}  # Objective column gets a higher cap; default cap is 28.
    for col_cells in ws.columns:
        col = col_cells[0].column
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(longest + 2, caps.get(col, 28))


def export_progress(db: sqlite3.Connection, subject_id: str, reports_root: str,
                    today: str | None = None) -> Path:
    """Generate the workbook and save it under REPORTS_ROOT. Returns the file path."""
    today = today or date.today().isoformat()
    rows = fetch_progress(db, subject_id)
    wb = build_workbook(rows, subject_id)

    out_dir = Path(reports_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{subject_id}_progress_{today}.xlsx"
    wb.save(out_path)
    return out_path


def main() -> None:
    ap = argparse.ArgumentParser(description="Export a subject's study progress to Excel.")
    ap.add_argument("--subject", required=True, help="e.g. Principles_of_Business")
    args = ap.parse_args()

    db_path = os.getenv("DB_PATH")
    reports_root = os.getenv("REPORTS_ROOT")
    if not db_path:
        sys.exit("ERROR: DB_PATH not set in .env")
    if not reports_root:
        sys.exit("ERROR: REPORTS_ROOT not set in .env")
    if not Path(db_path).exists():
        sys.exit(f"ERROR: database not found at {db_path}. Run init_db.py first.")

    db = open_db(db_path)
    try:
        rows = fetch_progress(db, args.subject)
        if not rows:
            sys.exit(f"ERROR: no objectives found for subject '{args.subject}'.")
        out_path = export_progress(db, args.subject, reports_root)
    finally:
        db.close()

    summary, mastered, total, _ = _summary_text(rows)
    print(f"{summary}")
    print(f"Progress report written: {out_path}")


if __name__ == "__main__":
    main()
