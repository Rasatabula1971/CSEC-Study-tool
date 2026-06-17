# PHASE: build
"""
backend/feedback_report.py
==========================
Stage 12 (Build Playbook v3.1) -- targeted teacher-review report.

The student taps 👍/👎/🤔 after each lesson and graded answer (logged in
user_feedback). This script turns that stream into a short, focused list: the
top 20 objectives by (negative + confused) feedback. The point is to convert an
impossible review task -- "check all 116 objectives" -- into a tractable one:
"these 5-20 are the ones a teacher should look at."

Everything here is deterministic SQLite + openpyxl (CLAUDE.md "Rule 2"): the
ranking comes straight from a GROUP BY on user_feedback, never the LLM. CASE WHEN
is used (not FILTER) for broader SQLite-version compatibility.

Usage:
    python backend/feedback_report.py --subject Principles_of_Business
"""

import argparse
import os
import sqlite3
import sys
from datetime import date
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(dotenv_path=Path(__file__).resolve().parents[1] / ".env")

SHEET_TITLE = "Top objectives for review"
HEADER_FILL = "2E75B6"  # blue header band
HEADERS = [
    "Objective ID", "Objective #", "Content (first 80 chars)",
    "Negative", "Confused", "Positive", "Total", "Last negative",
]
WIDTH_CAP = 60  # autofit ceiling per column


def open_db(db_path: str) -> sqlite3.Connection:
    try:
        import sqlite_vec
    except ImportError:
        sys.exit("ERROR: sqlite-vec is not installed. Run: pip install sqlite-vec")
    db = sqlite3.connect(db_path)
    db.enable_load_extension(True)
    sqlite_vec.load(db)
    db.enable_load_extension(False)
    db.execute("PRAGMA foreign_keys = ON")
    db.row_factory = sqlite3.Row
    return db


def fetch_flagged_objectives(db: sqlite3.Connection, subject_id: str) -> list[sqlite3.Row]:
    """Top 20 objectives by (negative + confused) feedback, descending.

    HAVING (neg + confused) > 0 means an objective that only got 👍 never appears --
    the report is the worry-list, not a full roster. CASE WHEN, not FILTER, so it
    runs on older SQLite builds.
    """
    return db.execute(
        """
        SELECT o.objective_id, o.objective_num,
               substr(o.content_stmt, 1, 80) AS content_short,
               COUNT(CASE WHEN f.sentiment='negative' THEN 1 END)  AS neg_count,
               COUNT(CASE WHEN f.sentiment='confused' THEN 1 END)  AS confused_count,
               COUNT(CASE WHEN f.sentiment='positive' THEN 1 END)  AS pos_count,
               COUNT(*)                                             AS total_feedback,
               MAX(CASE WHEN f.sentiment IN ('negative','confused')
                        THEN f.created_at END)                      AS last_negative
        FROM   user_feedback f
        JOIN   objectives o ON o.objective_id = f.objective_id
        WHERE  f.subject_id = ?
        GROUP  BY o.objective_id
        HAVING (neg_count + confused_count) > 0
        ORDER  BY (neg_count + confused_count) DESC,
                  total_feedback DESC
        LIMIT  20
        """,
        (subject_id,),
    ).fetchall()


def build_workbook(rows: list[sqlite3.Row]):
    """Assemble the workbook: bold blue frozen header (row 1), one row per objective.

    Pure openpyxl -- no DB, no file I/O. With zero rows it still writes the header,
    so the output path always exists (TASK 3d).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE

    # Row 1: header band -- bold white text on the blue fill.
    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Rows 2+: one per flagged objective, in the column order A..H.
    for r in rows:
        ws.append([
            r["objective_id"],
            r["objective_num"],
            r["content_short"],
            r["neg_count"],
            r["confused_count"],
            r["pos_count"],
            r["total_feedback"],
            r["last_negative"] or "",
        ])

    ws.freeze_panes = "A2"  # keep the header visible while scrolling
    _auto_width(ws)
    return wb


def _auto_width(ws) -> None:
    """Size each column to its widest cell (header included), capped at WIDTH_CAP."""
    for col_cells in ws.columns:
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(longest + 2, WIDTH_CAP)


def generate_report(db: sqlite3.Connection, subject_id: str, reports_root: str,
                    today: str | None = None) -> tuple[Path, int]:
    """Build the workbook and save it under reports_root. Returns (path, row_count).

    The directory is created if absent. An empty result still writes a
    header-only file so the path always exists.
    """
    today = today or date.today().isoformat()
    rows = fetch_flagged_objectives(db, subject_id)
    wb = build_workbook(rows)

    out_dir = Path(reports_root)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{subject_id}_feedback_report_{today}.xlsx"
    wb.save(out_path)
    return out_path, len(rows)


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Report the top objectives flagged by student feedback."
    )
    ap.add_argument("--subject", required=True, help="e.g. Principles_of_Business")
    args = ap.parse_args()

    db_path = os.getenv("DB_PATH")
    reports_root = os.getenv("REPORTS_ROOT")
    if not db_path:
        sys.exit("ERROR: DB_PATH not set in .env")
    if not reports_root:
        sys.exit("ERROR: REPORTS_ROOT not set in .env")
    if not Path(db_path).exists():
        sys.exit(f"ERROR: database not found at {db_path}. Run init_db.py first.")

    db = open_db(db_path)
    try:
        out_path, count = generate_report(db, args.subject, reports_root)
    finally:
        db.close()

    if count == 0:
        print(f"No feedback recorded yet for {args.subject}")
    print(f"Rows written: {count}")
    print(f"Feedback report written: {out_path}")


if __name__ == "__main__":
    main()
